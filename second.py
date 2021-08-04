import openpyxl
import requests
import bs4
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from bs4 import BeautifulSoup


#******* AVOID BEING BLOCKED *******
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


#******* WORKING WITH EXCEL *******
chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myAddresses.xlsx"
choose_SHEET_Of_Your_Excel_File = 'Hoja1'

#READER OF EXCEL FILE
wb = load_workbook(chooseExcel_File, data_only=True)
sh = wb[choose_SHEET_Of_Your_Excel_File]

#ADD Company Name from Excel
companyName = sh["c2"].value
book = Workbook()
ws = book.active

# companyName = input("What's the Company's name? ")#Add name manually
myAddress = input("Where is located the Company? ")#Add Address manually

#CREATING URL FOR GOOGLE SEARCHING
text = ("phone number of {0} {1} ".format(companyName, myAddress))
myCurrentURL = 'https://google.com/search?q=' + text

#ACCESING TO URL's INFO
response = requests.get(myCurrentURL, headers=headers)
soup = BeautifulSoup(response.text, 'html.parser')

tagSearched = soup.find_all('span', attrs={'class': 'mw31Ze'})
for tag in tagSearched:
    companyPhone = tag.text.strip()
    sh['E2'] = companyPhone
    wb.save(filename = chooseExcel_File)